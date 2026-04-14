from typing import List, Optional, Dict, Any
from pathlib import Path
import subprocess
import shutil
import pypdf
import docx2txt
from app.core.logger import get_logger
from langchain_text_splitters import (
    RecursiveCharacterTextSplitter,
    MarkdownHeaderTextSplitter,
    TokenTextSplitter,
    Language
)
from langchain_core.documents import Document
import jieba

logger = get_logger("file_processor")


class AdvancedTextSplitter:
    """高级文本分块器 - 支持多种分块策略"""
    
    @staticmethod
    def recursive_split(
        text: str,
        chunk_size: int = 500,
        chunk_overlap: int = 50,
        separators: List[str] = None
    ) -> List[str]:
        """
        递归字符分块（适合通用文本）
        
        Args:
            text: 待分割文本
            chunk_size: 块大小
            chunk_overlap: 块重叠大小
            separators: 分隔符列表
            
        Returns:
            文本块列表
        """
        if separators is None:
            separators = ["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
        
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
            separators=separators
        )
        
        return splitter.split_text(text)
    
    @staticmethod
    def semantic_split(
        text: str,
        embeddings = None,
        breakpoint_threshold_type: str = "percentile",
        breakpoint_threshold_amount: int = 95
    ) -> List[str]:
        """
        语义分块（基于内容含义分段）
        
        Args:
            text: 待分割文本
            embeddings: 嵌入模型
            breakpoint_threshold_type: 断点阈值类型
            breakpoint_threshold_amount: 阈值
            
        Returns:
            文本块列表
        """
        try:
            from langchain_experimental.text_splitter import SemanticChunker
            
            if embeddings:
                splitter = SemanticChunker(
                    embeddings=embeddings,
                    breakpoint_threshold_type=breakpoint_threshold_type,
                    breakpoint_threshold_amount=breakpoint_threshold_amount
                )
                return splitter.split_text(text)
        except ImportError:
            print("langchain_experimental 未安装，使用递归分块作为后备")
        
        return AdvancedTextSplitter.recursive_split(text)
    
    @staticmethod
    def markdown_split(text: str) -> List[str]:
        """
        Markdown标题分块（适合结构化文档）
        
        Args:
            text: Markdown文本
            
        Returns:
            文本块列表
        """
        headers_to_split_on = [
            ("#", "Header 1"),
            ("##", "Header 2"),
            ("###", "Header 3"),
            ("####", "Header 4"),
        ]
        
        splitter = MarkdownHeaderTextSplitter(
            headers_to_split_on=headers_to_split_on,
            strip_headers=False
        )
        
        return splitter.split_text(text)
    
    @staticmethod
    def token_split(
        text: str,
        chunk_size: int = 512,
        chunk_overlap: int = 50
    ) -> List[str]:
        """
        Token分块（基于token数量，适合LLM处理）
        
        Args:
            text: 待分割文本
            chunk_size: 块大小（token数）
            chunk_overlap: 块重叠大小
            
        Returns:
            文本块列表
        """
        splitter = TokenTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            encoding_name="cl100k_base"
        )
        
        return splitter.split_text(text)
    
    @staticmethod
    def legal_split(text: str, chunk_size: int = 800) -> List[Dict[str, Any]]:
        """
        法律文档专用分块（保留条款编号和上下文）
        
        Args:
            text: 法律文本
            chunk_size: 目标块大小
            
        Returns:
            包含块内容和元数据的列表
        """
        import re
        
        chunks = []
        
        article_pattern = r'(?:第[一二三四五六七八九十百千\d]+条|article\s*\d+|第\d+条)'
        clause_pattern = r'((?:一|二|三|四|五|六|七|八|九|十|①|②|③|④|⑤|⑥|⑦|⑧|⑨|⑩|\d+)[、.])'
        
        articles = re.split(article_pattern, text)
        
        if len(articles) > 1:
            for i in range(1, len(articles), 2):
                if i < len(articles):
                    article_content = articles[i]
                    if i + 1 < len(articles):
                        article_content += articles[i + 1]
                    
                    if len(article_content) > chunk_size:
                        sub_chunks = AdvancedTextSplitter.recursive_split(
                            article_content,
                            chunk_size=chunk_size,
                            chunk_overlap=100
                        )
                        for sub_idx, sub_chunk in enumerate(sub_chunks):
                            chunks.append({
                                "content": sub_chunk,
                                "metadata": {
                                    "article_index": i,
                                    "sub_chunk_index": sub_idx,
                                    "type": "legal_article"
                                }
                            })
                    else:
                        chunks.append({
                            "content": article_content,
                            "metadata": {
                                "article_index": i,
                                "type": "legal_article"
                            }
                        })
        else:
            sub_chunks = AdvancedTextSplitter.recursive_split(
                text,
                chunk_size=chunk_size,
                chunk_overlap=100
            )
            for idx, chunk in enumerate(sub_chunks):
                chunks.append({
                    "content": chunk,
                    "metadata": {
                        "chunk_index": idx,
                        "type": "legal_general"
                    }
                })
        
        return chunks


class FileProcessor:
    """文件处理器，支持多种文件格式和高级分块"""
    
    def __init__(
        self,
        chunk_size: int = 500,
        chunk_overlap: int = 50,
        split_strategy: str = "recursive"
    ):
        """
        初始化文件处理器
        
        Args:
            chunk_size: 文本分块大小
            chunk_overlap: 文本分块重叠大小
            split_strategy: 分块策略 (recursive/markdown/token/legal)
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.split_strategy = split_strategy
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
        )
    
    def extract_text(self, file_path: str) -> str:
        """
        从文件中提取文本

        Args:
            file_path: 文件路径

        Returns:
            提取的文本内容
        """
        file_ext = Path(file_path).suffix.lower()
        logger.info(f"开始提取文件文本 - file_path: {file_path}, file_ext: {file_ext}")

        if file_ext == '.pdf':
            text = self._extract_pdf(file_path)
        elif file_ext in ['.docx', '.doc']:
            text = self._extract_docx(file_path)
        elif file_ext in ['.txt', '.md']:
            text = self._extract_text_file(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            text = self._extract_excel(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {file_ext}")

        logger.info(f"文件文本提取结束 - file_path: {file_path}, text_length: {len(text) if text else 0}")
        return text
    
    def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文件文本"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text += f"[页面{page_num + 1}]\n{page_text}\n"
        except Exception as e:
            print(f"提取PDF文本失败: {e}")
            return ""
        return text
    
    def _extract_docx(self, file_path: str) -> str:
        """提取DOC/DOCX文件文本"""
        file_ext = Path(file_path).suffix.lower()

        if file_ext == '.docx':
            try:
                text = docx2txt.process(file_path)
                if text:
                    logger.info(f"DOCX文本提取成功 - file_path: {file_path}, text_length: {len(text)}")
                    return text
            except Exception as e:
                logger.warning(f"DOCX文本提取失败 - file_path: {file_path}, error: {e}")
            return ""

        if file_ext == '.doc':
            text = self._extract_doc_with_win32com(file_path)
            if text:
                return text

            antiword_path = shutil.which('antiword')
            if antiword_path:
                try:
                    antiword_args = [antiword_path]
                    mapping_path = self._get_antiword_utf8_mapping(antiword_path)
                    if mapping_path:
                        antiword_args.extend(['-m', mapping_path])
                    antiword_args.append(file_path)

                    result = subprocess.run(
                        antiword_args,
                        capture_output=True,
                        check=False
                    )
                    text = self._decode_text_bytes(result.stdout).strip()
                    stderr_text = self._decode_text_bytes(result.stderr).strip()
                    if result.returncode == 0 and text:
                        logger.info(f"DOC文本提取成功(antiword) - file_path: {file_path}, text_length: {len(text)}")
                        return text
                    logger.warning(
                        f"DOC文本提取失败(antiword) - file_path: {file_path}, returncode: {result.returncode}, stderr: {stderr_text[:200]}"
                    )
                except Exception as e:
                    logger.warning(f"DOC文本提取异常(antiword) - file_path: {file_path}, error: {e}")

            try:
                import textract
                text = self._decode_text_bytes(textract.process(file_path)).strip()
                if text:
                    logger.info(f"DOC文本提取成功(textract) - file_path: {file_path}, text_length: {len(text)}")
                    return text
            except Exception as e:
                logger.warning(f"DOC文本提取失败(textract) - file_path: {file_path}, error: {e}")

        return ""

    def _extract_doc_with_win32com(self, file_path: str) -> str:
        """使用 win32com 提取 .doc 文件文本 (Windows 专用)"""
        import platform
        if platform.system() != 'Windows':
            return ""

        try:
            import win32com.client
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(Path(file_path).resolve()))
            text = doc.Content.Text
            doc.Close(False)
            word.Quit()
            if text:
                logger.info(f"DOC文本提取成功(win32com) - file_path: {file_path}, text_length: {len(text)}")
                return text
        except ImportError:
            logger.warning("win32com 未安装，无法使用 Word COM 接口提取 .doc 文件")
        except Exception as e:
            logger.warning(f"DOC文本提取失败(win32com) - file_path: {file_path}, error: {e}")

        return ""

    def _get_antiword_utf8_mapping(self, antiword_path: str) -> str | None:
        antiword_dir = Path(antiword_path).resolve().parent
        candidates = [
            antiword_dir.parent / 'share' / 'antiword' / 'UTF-8.txt',
            antiword_dir / 'UTF-8.txt',
            Path('UTF-8.txt')
        ]

        for candidate in candidates:
            if candidate.exists():
                return str(candidate)

        return None

    def _decode_text_bytes(self, content: bytes | str | None) -> str:
        if content is None:
            return ""
        if isinstance(content, str):
            return content

        for encoding in ('utf-8', 'gb18030', 'gbk', 'cp936', 'latin1'):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue

        return content.decode('utf-8', errors='ignore')
    
    def _extract_text_file(self, file_path: str) -> str:
        """提取文本文件内容"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    return file.read()
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"提取文本文件失败: {e}")
                return ""
        
        return ""
    
    def _extract_excel(self, file_path: str) -> str:
        """提取Excel文件文本"""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            text = ""
            for sheet in workbook:
                text += f"[工作表: {sheet.title}]\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        except Exception as e:
            print(f"提取Excel文本失败: {e}")
            return ""
    
    def split_text(self, text: str, strategy: str = None) -> List[str]:
        """
        将文本分割成块
        
        Args:
            text: 待分割的文本
            strategy: 分块策略（默认使用实例配置）
            
        Returns:
            文本块列表
        """
        if not text or len(text.strip()) == 0:
            return []
        
        strategy = strategy or self.split_strategy
        
        if strategy == "recursive":
            return self.text_splitter.split_text(text)
        elif strategy == "markdown":
            return AdvancedTextSplitter.markdown_split(text)
        elif strategy == "token":
            return AdvancedTextSplitter.token_split(text)
        elif strategy == "legal":
            legal_chunks = AdvancedTextSplitter.legal_split(text, self.chunk_size)
            return [chunk["content"] for chunk in legal_chunks]
        else:
            return self.text_splitter.split_text(text)
    
    def split_text_with_metadata(
        self,
        text: str,
        strategy: str = None
    ) -> List[Document]:
        """
        将文本分割成块并保留元数据
        
        Args:
            text: 待分割的文本
            strategy: 分块策略
            
        Returns:
            Document列表
        """
        chunks = self.split_text(text, strategy)
        
        documents = []
        for i, chunk in enumerate(chunks):
            documents.append(Document(
                page_content=chunk,
                metadata={
                    "chunk_index": i,
                    "total_chunks": len(chunks),
                    "chunk_size": len(chunk),
                    "split_strategy": strategy or self.split_strategy
                }
            ))
        
        return documents
    
    def process_file(
        self,
        file_path: str,
        metadata: dict = None,
        strategy: str = None
    ) -> tuple[List[str], List[dict]]:
        """
        处理文件：提取文本并分割成块
        
        Args:
            file_path: 文件路径
            metadata: 基础元数据
            strategy: 分块策略
            
        Returns:
            (文本块列表, 元数据列表)
        """
        text = self.extract_text(file_path)
        if not text:
            return [], []
        
        chunks = self.split_text(text, strategy)
        
        if metadata is None:
            metadata = {}
        
        metadatas = []
        for i, chunk in enumerate(chunks):
            chunk_metadata = metadata.copy()
            chunk_metadata.update({
                "chunk_index": i,
                "total_chunks": len(chunks),
                "chunk_size": len(chunk),
                "split_strategy": strategy or self.split_strategy
            })
            metadatas.append(chunk_metadata)
        
        return chunks, metadatas
    
    def process_file_with_documents(
        self,
        file_path: str,
        metadata: dict = None,
        strategy: str = None
    ) -> List[Document]:
        """
        处理文件并返回Document列表
        
        Args:
            file_path: 文件路径
            metadata: 基础元数据
            strategy: 分块策略
            
        Returns:
            Document列表
        """
        text = self.extract_text(file_path)
        if not text:
            return []
        
        chunks = self.split_text(text, strategy)
        
        if metadata is None:
            metadata = {}
        
        documents = []
        for i, chunk in enumerate(chunks):
            chunk_metadata = metadata.copy()
            chunk_metadata.update({
                "chunk_index": i,
                "total_chunks": len(chunks),
                "chunk_size": len(chunk),
                "split_strategy": strategy or self.split_strategy
            })
            documents.append(Document(
                page_content=chunk,
                metadata=chunk_metadata
            ))
        
        return documents
